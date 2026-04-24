import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def processar_excel_robusto(nome_arquivo):
    print("Lendo o arquivo Excel... (isso pode levar alguns segundos)")
    try:
        wb = load_workbook(nome_arquivo)
    except Exception as e:
        print(f"Erro ao abrir o arquivo: {e}")
        return

    vermelho = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    amarelo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"Processando a aba: {sheet_name}")

        produto_atual_linha = None
        lotes_do_produto = {}

        for linha_idx, row in enumerate(ws.iter_rows(), start=1):
            # Junta todo o texto da linha para investigar, não importa em qual coluna esteja
            texto_linha = " ".join([str(cell.value).strip() for cell in row if cell.value is not None])
            
            if not texto_linha or 'DESCRIÇÃO DOS PRODUTOS' in texto_linha.upper():
                continue

            # A MÁGICA: Procura por um padrão de data (ex: 30/12/30 ou 30/12/2030) na linha
            # Se tiver data, temos 100% de certeza que é uma linha de LOTE
            if re.search(r'\d{2}/\d{2}/\d{2,4}', texto_linha):
                if produto_atual_linha is not None:
                    # O código do lote é o primeiro bloquinho de texto da célula
                    codigo_lote_original = ""
                    for cell in row:
                        if cell.value:
                            codigo_lote_original = str(cell.value).strip().split()[0]
                            break
                    
                    # Limpa para achar os erros de digitação (amarelo)
                    codigo_limpo = re.sub(r'[\s-]', '', codigo_lote_original).upper()

                    if not codigo_limpo:
                        continue

                    # Compara com os lotes anteriores daquele mesmo produto
                    for lote_salvo, lote_limpo_salvo in lotes_do_produto.items():
                        if codigo_lote_original == lote_salvo:
                            # Igual exato: pinta o PRODUTO de vermelho
                            ws.cell(row=produto_atual_linha, column=1).fill = vermelho
                        elif codigo_limpo == lote_limpo_salvo:
                            # Quase igual: pinta a linha do LOTE de amarelo
                            ws.cell(row=linha_idx, column=1).fill = amarelo

                    # Salva o lote para continuar a comparação
                    lotes_do_produto[codigo_lote_original] = codigo_limpo

            else:
                # Se NÃO tem data e a linha começa com um número (código do produto)
                primeiro_valor = str(row[0].value).strip() if row[0].value else ""
                if primeiro_valor and primeiro_valor[0].isdigit():
                    produto_atual_linha = linha_idx
                    lotes_do_produto = {} # Zera os lotes porque é um produto novo

    novo_nome = 'LOTES_ORGANIZADOS_CORRIGIDO.xlsx'
    wb.save(novo_nome)
    print(f"\nPronto! O novo arquivo foi salvo como: {novo_nome}")

# Executa o código
processar_excel_robusto('LOTES-E-PRODUTOS_FELIPE23_.xlsx')